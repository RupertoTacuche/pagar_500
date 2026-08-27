from fastapi import FastAPI, Form, Request
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.templating import Jinja2Templates
import openpyxl
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

app = FastAPI()
templates = Jinja2Templates(directory="templates")

# Permitir CORS para que el formulario HTML pueda enviar datos
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def formulario(request: Request):
    return templates.TemplateResponse(request=request, name="formulario.html")


@app.post("/guardar-cliente")
async def guardar_cliente(
    nombre: str = Form(...),
    email: str = Form(...),
    servicio: str = Form(...)
):
    archivo = "clientes.xlsx"

    if os.path.exists(archivo):
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nombre", "Email", "Servicio"])

    ws.append([nombre, email, servicio])
    wb.save(archivo)

    return JSONResponse({"mensaje": "Cliente guardado", "nombre": nombre})


@app.get("/generar-reporte")
def generar_reporte():
    archivo_excel = "clientes.xlsx"
    archivo_pdf = "reporte_clientes.pdf"

    wb = openpyxl.load_workbook(archivo_excel)
    ws = wb.active

    c = canvas.Canvas(archivo_pdf, pagesize=letter)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, 750, "Reporte de Clientes")

    c.setFont("Helvetica", 12)
    y = 720

    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre, email, servicio = fila
        c.drawString(50, y, f"{nombre} | {email} | {servicio}")
        y -= 20

        if y < 50:
            c.showPage()
            y = 750

    c.save()
    return FileResponse(archivo_pdf, media_type="application/pdf")
